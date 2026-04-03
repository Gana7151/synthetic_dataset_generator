"""
Generates supporting input documents:
  - W-2 (PDF)
  - 1099-INT (PDF)
  - 1099-DIV (PDF)
  - Bank Statement (PDF + XLSX)
  - Invoice (PDF)
  - Schedule C Expense Ledger (XLSX)
"""

import os
import random
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                 TableStyle, HRFlowable, PageBreak)
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from generators.pdf_styles import (
    get_styles, format_ssn, format_ein, format_currency, format_currency_int,
    DARK_BLUE, MEDIUM_BLUE, LIGHT_BLUE, HEADER_BG,
    BORDER_COLOR, WHITE, LIGHT_GRAY, BLACK, DARK_GRAY, MEDIUM_GRAY,
    GREEN, RED, MARGIN, PAGE_WIDTH, PAGE_HEIGHT
)


def generate_all_input_documents(profile, output_dir: str):
    """Generate all input documents for a profile into output_dir."""
    os.makedirs(output_dir, exist_ok=True)

    # W-2 PDFs
    for i, w2 in enumerate(profile.w2_incomes):
        suffix = f"_{i+1}" if len(profile.w2_incomes) > 1 else ""
        _generate_w2_pdf(w2, profile,
                         os.path.join(output_dir, f"Form_W-2{suffix}.pdf"))

    # 1099-INT PDFs
    for i, interest in enumerate(profile.interest_incomes):
        suffix = f"_{i+1}" if len(profile.interest_incomes) > 1 else ""
        _generate_1099_int_pdf(interest, profile,
                               os.path.join(output_dir, f"1099-INT{suffix}.pdf"))

    # 1099-DIV PDFs
    for i, div in enumerate(profile.dividend_incomes):
        suffix = f"_{i+1}" if len(profile.dividend_incomes) > 1 else ""
        _generate_1099_div_pdf(div, profile,
                               os.path.join(output_dir, f"1099-DIV{suffix}.pdf"))

    # 1098 PDFs
    for i, m in enumerate(getattr(profile, 'mortgage_interests', [])):
        suffix = f"_{i+1}" if len(getattr(profile, 'mortgage_interests', [])) > 1 else ""
        _generate_1098_pdf(m, profile, os.path.join(output_dir, f"1098{suffix}.pdf"))

    # 1099-B PDFs
    for i, cg in enumerate(getattr(profile, 'capital_gains', [])):
        suffix = f"_{i+1}" if len(getattr(profile, 'capital_gains', [])) > 1 else ""
        _generate_1099_b_pdf(cg, profile, os.path.join(output_dir, f"1099-B{suffix}.pdf"))

    # Bank Statement (if business income exists)
    if profile.business_income:
        _generate_bank_statement_xlsx(profile,
            os.path.join(output_dir, f"Bank_Statement_{profile.tax_year}.xlsx"))
        _generate_invoice_pdf(profile,
            os.path.join(output_dir, "Invoice.pdf"))
        _generate_schedule_c_xlsx(profile,
            os.path.join(output_dir, f"Schedule_C_Ledger_{profile.tax_year}.xlsx"))

    # Tax Document Attachments Summary
    _generate_attachments_summary(profile, output_dir)


# ===========================================================================
# W-2 PDF Generator
# ===========================================================================

def _generate_w2_pdf(w2, profile, output_path: str):
    """Generate a W-2 form PDF."""
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter

    # Title bar
    c.setFillColor(DARK_BLUE)
    c.rect(0.5*inch, h - 1.0*inch, w - 1.0*inch, 0.4*inch, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(w/2, h - 0.82*inch, f"Form W-2  —  Wage and Tax Statement  ({profile.tax_year})")

    # Subtitle
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w/2, h - 1.15*inch, "Department of the Treasury — Internal Revenue Service")

    y = h - 1.6*inch

    def draw_box(x, box_y, box_w, box_h, label, value, label_prefix=""):
        c.setStrokeColor(BORDER_COLOR)
        c.setLineWidth(0.5)
        c.rect(x, box_y, box_w, box_h, stroke=1, fill=0)
        c.setFillColor(MEDIUM_GRAY)
        c.setFont("Helvetica", 6)
        c.drawString(x + 4, box_y + box_h - 8, f"{label_prefix}{label}")
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 4, box_y + 6, str(value))

    left = 0.5 * inch
    mid = w / 2
    box_w = (w - 1.2*inch) / 2
    box_h = 0.55 * inch

    # Row 1: Employee SSN | Employer EIN
    draw_box(left, y, box_w, box_h, "Employee's SSN", format_ssn(w2.employee_ssn), "a  ")
    draw_box(left + box_w + 0.2*inch, y, box_w, box_h, "Employer's EIN",
             format_ein(w2.employer_ein), "b  ")
    y -= box_h + 0.1*inch

    # Row 2: Employer Name & Address | Control Number
    big_box_h = 0.8 * inch
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, box_w, big_box_h, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + big_box_h - 8, "c  Employer's name, address, and ZIP code")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left + 4, y + big_box_h - 22, w2.employer_name)
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + big_box_h - 34, w2.employer_address)
    c.drawString(left + 4, y + big_box_h - 46,
                 f"{w2.employer_city}, {w2.employer_state} {w2.employer_zip}")

    draw_box(left + box_w + 0.2*inch, y + big_box_h - box_h, box_w, box_h,
             "Control number", "", "d  ")
    y -= big_box_h + 0.1*inch

    # Row 3: Employee name & address
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, box_w, big_box_h, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + big_box_h - 8, "e/f  Employee's name and address")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(left + 4, y + big_box_h - 22, w2.employee_name)
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + big_box_h - 34, profile.address)
    c.drawString(left + 4, y + big_box_h - 46,
                 f"{profile.city}, {profile.state} {profile.zip_code}")

    y -= big_box_h + 0.15*inch

    # Wage boxes (2 columns)
    wage_boxes = [
        ("1  Wages, tips, other compensation", format_currency(w2.wages)),
        ("2  Federal income tax withheld", format_currency(w2.federal_withheld)),
        ("3  Social security wages", format_currency(w2.ss_wages)),
        ("4  Social security tax withheld", format_currency(w2.ss_tax)),
        ("5  Medicare wages and tips", format_currency(w2.medicare_wages)),
        ("6  Medicare tax withheld", format_currency(w2.medicare_tax)),
        ("7  Social security tips", format_currency(0)),
        ("8  Allocated tips", format_currency(0)),
    ]

    small_box_h = 0.45 * inch
    for i in range(0, len(wage_boxes), 2):
        label1, val1 = wage_boxes[i]
        c.setStrokeColor(BORDER_COLOR)
        c.rect(left, y, box_w, small_box_h, stroke=1, fill=0)
        c.setFillColor(MEDIUM_GRAY)
        c.setFont("Helvetica", 6)
        c.drawString(left + 4, y + small_box_h - 8, label1)
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(left + 4, y + 6, val1)

        if i + 1 < len(wage_boxes):
            label2, val2 = wage_boxes[i + 1]
            c.setStrokeColor(BORDER_COLOR)
            c.rect(left + box_w + 0.2*inch, y, box_w, small_box_h, stroke=1, fill=0)
            c.setFillColor(MEDIUM_GRAY)
            c.setFont("Helvetica", 6)
            c.drawString(left + box_w + 0.24*inch, y + small_box_h - 8, label2)
            c.setFillColor(BLACK)
            c.setFont("Helvetica-Bold", 11)
            c.drawString(left + box_w + 0.24*inch, y + 6, val2)

        y -= small_box_h + 0.05*inch

    # State/local section
    y -= 0.1*inch
    state_box_h = 0.45 * inch
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, state_box_h, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + state_box_h - 8,
                 "15  State | Employer's state ID | 16  State wages | 17  State income tax")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 9)
    state_info = (f"{profile.state}        EIN: {format_ein(w2.employer_ein)}"
                  f"        {format_currency(w2.state_wages)}"
                  f"        {format_currency(w2.state_withheld)}")
    c.drawString(left + 4, y + 8, state_info)

    c.save()


# ===========================================================================
# 1099-INT PDF Generator
# ===========================================================================

def _generate_1099_int_pdf(interest, profile, output_path: str):
    """Generate a 1099-INT form PDF."""
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter

    # Title
    c.setFillColor(DARK_BLUE)
    c.rect(0.5*inch, h - 1.0*inch, w - 1.0*inch, 0.4*inch, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w/2, h - 0.82*inch,
                        f"Form 1099-INT  —  Interest Income  ({profile.tax_year})")

    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w/2, h - 1.15*inch,
                        "Department of the Treasury — Internal Revenue Service")

    y = h - 1.6*inch
    left = 0.5 * inch
    box_w = (w - 1.2*inch) / 2
    box_h = 0.6 * inch

    # Payer info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "PAYER'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24, interest.payer_name)
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"EIN: {format_ein(interest.payer_ein)}")
    y -= 0.9*inch

    # Recipient info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "RECIPIENT'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24,
                 f"{profile.primary_first} {profile.primary_last}")
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"SSN: {format_ssn(profile.primary_ssn)}")
    c.drawString(left + 4, y + 0.8*inch - 50,
                 f"{profile.address}, {profile.city}, {profile.state} {profile.zip_code}")
    y -= 0.9*inch

    # Amount boxes
    boxes = [
        ("1  Interest income", format_currency(interest.amount)),
        ("2  Early withdrawal penalty", format_currency(0)),
        ("3  Interest on U.S. Savings Bonds", format_currency(0)),
        ("4  Federal income tax withheld", format_currency(0)),
    ]

    for label, val in boxes:
        c.setStrokeColor(BORDER_COLOR)
        c.rect(left, y, w - 1.0*inch, box_h, stroke=1, fill=0)
        c.setFillColor(MEDIUM_GRAY)
        c.setFont("Helvetica", 7)
        c.drawString(left + 4, y + box_h - 10, label)
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(left + 4, y + 8, val)
        y -= box_h + 0.05*inch

    c.save()


# ===========================================================================
# 1099-DIV PDF Generator
# ===========================================================================

def _generate_1099_div_pdf(div, profile, output_path: str):
    """Generate a 1099-DIV form PDF."""
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter

    # Title
    c.setFillColor(DARK_BLUE)
    c.rect(0.5*inch, h - 1.0*inch, w - 1.0*inch, 0.4*inch, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w/2, h - 0.82*inch,
                        f"Form 1099-DIV  —  Dividends and Distributions  ({profile.tax_year})")

    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w/2, h - 1.15*inch,
                        "Department of the Treasury — Internal Revenue Service")

    y = h - 1.6*inch
    left = 0.5 * inch
    box_h = 0.55 * inch

    # Payer info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "PAYER'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24, div.payer_name)
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"EIN: {format_ein(div.payer_ein)}")
    y -= 0.9*inch

    # Recipient info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "RECIPIENT'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24,
                 f"{profile.primary_first} {profile.primary_last}")
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"SSN: {format_ssn(profile.primary_ssn)}")
    c.drawString(left + 4, y + 0.8*inch - 50,
                 f"{profile.address}, {profile.city}, {profile.state} {profile.zip_code}")
    y -= 0.9*inch

    # Dividend boxes
    boxes = [
        ("1a  Total ordinary dividends", format_currency(div.ordinary_dividends)),
        ("1b  Qualified dividends", format_currency(div.qualified_dividends)),
        ("2a  Total capital gain distr.", format_currency(0)),
        ("3  Nondividend distributions", format_currency(0)),
        ("4  Federal income tax withheld", format_currency(0)),
    ]

    for label, val in boxes:
        c.setStrokeColor(BORDER_COLOR)
        c.rect(left, y, w - 1.0*inch, box_h, stroke=1, fill=0)
        c.setFillColor(MEDIUM_GRAY)
        c.setFont("Helvetica", 7)
        c.drawString(left + 4, y + box_h - 10, label)
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(left + 4, y + 8, val)
        y -= box_h + 0.05*inch

    c.save()


# ===========================================================================
# 1098 PDF Generator
# ===========================================================================

def _generate_1098_pdf(mortgage, profile, output_path: str):
    """Generate a 1098 form PDF."""
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter

    # Title
    c.setFillColor(DARK_BLUE)
    c.rect(0.5*inch, h - 1.0*inch, w - 1.0*inch, 0.4*inch, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w/2, h - 0.82*inch,
                        f"Form 1098  —  Mortgage Interest Statement  ({profile.tax_year})")

    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w/2, h - 1.15*inch,
                        "Department of the Treasury — Internal Revenue Service")

    y = h - 1.6*inch
    left = 0.5 * inch
    box_h = 0.55 * inch

    # Payer info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "LENDER'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24, mortgage.lender_name)
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"EIN: {format_ein(mortgage.lender_ein)}")
    y -= 0.9*inch

    # Recipient info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "PAYER'S/BORROWER'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24, f"{profile.primary_first} {profile.primary_last}")
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"SSN: {format_ssn(profile.primary_ssn)}")
    c.drawString(left + 4, y + 0.8*inch - 50, f"{profile.address}, {profile.city}, {profile.state} {profile.zip_code}")
    y -= 0.9*inch

    # Boxes
    boxes = [
        ("1  Mortgage interest received", format_currency(mortgage.interest_paid)),
        ("2  Outstanding mortgage principal", format_currency(mortgage.principal)),
        ("3  Mortgage origination date", "01/15/2015"),
        ("4  Refund of overpaid interest", format_currency(0)),
        ("5  Mortgage insurance premiums", format_currency(0)),
    ]

    for label, val in boxes:
        c.setStrokeColor(BORDER_COLOR)
        c.rect(left, y, w - 1.0*inch, box_h, stroke=1, fill=0)
        c.setFillColor(MEDIUM_GRAY)
        c.setFont("Helvetica", 7)
        c.drawString(left + 4, y + box_h - 10, label)
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(left + 4, y + 8, val)
        y -= box_h + 0.05*inch

    c.save()


# ===========================================================================
# 1099-B PDF Generator
# ===========================================================================

def _generate_1099_b_pdf(cg, profile, output_path: str):
    """Generate a 1099-B form PDF."""
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter

    # Title
    c.setFillColor(DARK_BLUE)
    c.rect(0.5*inch, h - 1.0*inch, w - 1.0*inch, 0.4*inch, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(w/2, h - 0.82*inch,
                        f"Form 1099-B  —  Proceeds From Broker  ({profile.tax_year})")

    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w/2, h - 1.15*inch,
                        "Department of the Treasury — Internal Revenue Service")

    y = h - 1.6*inch
    left = 0.5 * inch
    box_h = 0.55 * inch

    # Payer info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "PAYER'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24, cg.payer_name)
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"EIN: {format_ein(cg.payer_ein)}")
    y -= 0.9*inch

    # Recipient info
    c.setStrokeColor(BORDER_COLOR)
    c.rect(left, y, w - 1.0*inch, 0.8*inch, stroke=1, fill=0)
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 6)
    c.drawString(left + 4, y + 0.8*inch - 8, "RECIPIENT'S name, address, and TIN")
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(left + 4, y + 0.8*inch - 24, f"{profile.primary_first} {profile.primary_last}")
    c.setFont("Helvetica", 8)
    c.drawString(left + 4, y + 0.8*inch - 38, f"SSN: {format_ssn(profile.primary_ssn)}")
    c.drawString(left + 4, y + 0.8*inch - 50, f"{profile.address}, {profile.city}, {profile.state} {profile.zip_code}")
    y -= 0.9*inch

    # Boxes (assuming 20% gain margin for proceeds)
    st_proceeds = cg.short_term_gains * 5 if cg.short_term_gains > 0 else 0
    st_cost = st_proceeds - cg.short_term_gains

    lt_proceeds = cg.long_term_gains * 5 if cg.long_term_gains > 0 else 0
    lt_cost = lt_proceeds - cg.long_term_gains

    boxes = [
        ("1d  Proceeds (Short-Term)", format_currency(st_proceeds)),
        ("1e  Cost or other basis (Short-Term)", format_currency(st_cost)),
        ("1d  Proceeds (Long-Term)", format_currency(lt_proceeds)),
        ("1e  Cost or other basis (Long-Term)", format_currency(lt_cost)),
        ("4  Federal income tax withheld", format_currency(0)),
    ]

    for label, val in boxes:
        c.setStrokeColor(BORDER_COLOR)
        c.rect(left, y, w - 1.0*inch, box_h, stroke=1, fill=0)
        c.setFillColor(MEDIUM_GRAY)
        c.setFont("Helvetica", 7)
        c.drawString(left + 4, y + box_h - 10, label)
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(left + 4, y + 8, val)
        y -= box_h + 0.05*inch

    c.save()


# ===========================================================================
# Bank Statement (XLSX)
# ===========================================================================

def _generate_bank_statement_xlsx(profile, output_path: str):
    """Generate a bank statement Excel file."""
    biz = profile.business_income
    if not biz:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    currency_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = random.choice(["First National Bank", "Pacific Coast Credit Union",
                               "Citizens Trust Bank"])
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1a365d')

    ws.merge_cells('A2:F2')
    ws['A2'] = f"Business Account Statement — {profile.tax_year}"
    ws['A2'].font = Font(name='Calibri', size=11, italic=True)

    ws.merge_cells('A3:F3')
    ws['A3'] = f"Account Holder: {biz.business_name}"

    ws.merge_cells('A4:F4')
    ws['A4'] = f"Period: January 1 – December 31, {profile.tax_year}"

    # Headers
    row = 6
    headers = ["Date", "Description", "Type", "Debit", "Credit", "Balance"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Generate monthly transactions
    monthly_revenue = biz.gross_receipts / 12
    monthly_expenses = biz.expenses.total / 12
    balance = round(random.uniform(5000, 15000), 2)
    row += 1

    # Opening balance
    ws.cell(row=row, column=1, value=f"01/01/{profile.tax_year}")
    ws.cell(row=row, column=2, value="Opening Balance")
    ws.cell(row=row, column=3, value="—")
    ws.cell(row=row, column=6, value=balance).number_format = currency_fmt
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    expense_descs = [
        "Office Supplies - Staples", "Adobe Creative Cloud",
        "Internet Service Provider", "Phone Service", "Advertising - Google Ads",
        "Professional Insurance", "Fuel / Transportation", "Equipment Purchase",
        "Software License", "Marketing Materials",
    ]

    client_names = [
        "Acme Corp", "TechStart Inc.", "BlueSky Design", "Metro Brands",
        "Peak Performance LLC", "Greenfield Consulting", "Urban Solutions",
    ]

    for month in range(1, 13):
        # 2-4 revenue entries per month
        num_revenues = random.randint(2, 4)
        rev_per = monthly_revenue / num_revenues

        for j in range(num_revenues):
            day = random.randint(1, 28)
            amount = round(rev_per * random.uniform(0.7, 1.3), 2)
            balance = round(balance + amount, 2)
            ws.cell(row=row, column=1, value=f"{month:02d}/{day:02d}/{profile.tax_year}")
            ws.cell(row=row, column=2, value=f"Client Payment - {random.choice(client_names)}")
            ws.cell(row=row, column=3, value="Deposit")
            ws.cell(row=row, column=5, value=amount).number_format = currency_fmt
            ws.cell(row=row, column=6, value=balance).number_format = currency_fmt
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

        # 2-3 expense entries per month
        num_expenses = random.randint(2, 3)
        exp_per = monthly_expenses / num_expenses

        for j in range(num_expenses):
            day = random.randint(1, 28)
            amount = round(exp_per * random.uniform(0.6, 1.4), 2)
            balance = round(balance - amount, 2)
            ws.cell(row=row, column=1, value=f"{month:02d}/{day:02d}/{profile.tax_year}")
            ws.cell(row=row, column=2, value=random.choice(expense_descs))
            ws.cell(row=row, column=3, value="Withdrawal")
            ws.cell(row=row, column=4, value=amount).number_format = currency_fmt
            ws.cell(row=row, column=6, value=balance).number_format = currency_fmt
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
            row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    wb.save(output_path)


# ===========================================================================
# Invoice PDF Generator
# ===========================================================================

def _generate_invoice_pdf(profile, output_path: str):
    """Generate a sample business invoice PDF."""
    biz = profile.business_income
    c = canvas.Canvas(output_path, pagesize=letter)
    w, h = letter

    # Header
    c.setFillColor(DARK_BLUE)
    c.rect(0, h - 1.2*inch, w, 1.2*inch, fill=1, stroke=0)

    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 22)
    c.drawString(0.6*inch, h - 0.7*inch, biz.business_name)
    c.setFont("Helvetica", 10)
    c.drawString(0.6*inch, h - 0.95*inch,
                 f"{profile.address}, {profile.city}, {profile.state} {profile.zip_code}")

    c.setFont("Helvetica-Bold", 28)
    c.drawRightString(w - 0.6*inch, h - 0.75*inch, "INVOICE")

    # Invoice details
    y = h - 1.6*inch
    c.setFillColor(BLACK)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(0.6*inch, y, f"Invoice #: INV-{profile.tax_year}-{random.randint(1000,9999)}")
    c.drawString(0.6*inch, y - 16, f"Date: {random.randint(1,12):02d}/15/{profile.tax_year}")

    c.drawRightString(w - 0.6*inch, y, "Bill To:")
    c.setFont("Helvetica", 9)
    client = random.choice(["Acme Corporation", "TechStart Inc.",
                            "BlueSky Design Agency", "Metro Brands LLC"])
    c.drawRightString(w - 0.6*inch, y - 14, client)
    c.drawRightString(w - 0.6*inch, y - 26,
                      f"{random.randint(100,999)} {random.choice(['Main','Oak','Pine','Elm'])} Street")

    # Line items table
    y -= 0.8*inch
    # Table header
    c.setFillColor(HEADER_BG)
    c.rect(0.5*inch, y, w - 1.0*inch, 0.3*inch, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(0.6*inch, y + 8, "Description")
    c.drawString(4.5*inch, y + 8, "Qty")
    c.drawString(5.3*inch, y + 8, "Rate")
    c.drawRightString(w - 0.6*inch, y + 8, "Amount")

    y -= 0.05*inch
    c.setFillColor(BLACK)
    c.setFont("Helvetica", 9)

    services = [
        (f"{biz.activity_desc} — Monthly retainer", 1,
         round(biz.gross_receipts * 0.4)),
        (f"{biz.activity_desc} — Project deliverable", 1,
         round(biz.gross_receipts * 0.35)),
        ("Additional consultation hours", random.randint(5, 20),
         round(random.uniform(75, 150))),
    ]

    total = 0
    for desc, qty, rate in services:
        y -= 0.3*inch
        amount = qty * rate
        total += amount
        c.drawString(0.6*inch, y + 8, desc)
        c.drawString(4.5*inch, y + 8, str(qty))
        c.drawString(5.3*inch, y + 8, format_currency(rate))
        c.drawRightString(w - 0.6*inch, y + 8, format_currency(amount))
        c.setStrokeColor(LIGHT_GRAY)
        c.line(0.5*inch, y, w - 0.5*inch, y)

    # Total
    y -= 0.5*inch
    c.setStrokeColor(DARK_BLUE)
    c.setLineWidth(2)
    c.line(4.5*inch, y + 0.35*inch, w - 0.5*inch, y + 0.35*inch)
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(DARK_BLUE)
    c.drawString(5.3*inch, y + 10, "TOTAL:")
    c.drawRightString(w - 0.6*inch, y + 10, format_currency(total))

    # Footer
    c.setFillColor(MEDIUM_GRAY)
    c.setFont("Helvetica", 8)
    c.drawCentredString(w/2, 0.5*inch, "Thank you for your business!")

    c.save()


# ===========================================================================
# Schedule C Expense Ledger (XLSX)
# ===========================================================================

def _generate_schedule_c_xlsx(profile, output_path: str):
    """Generate a Schedule C expense breakdown Excel file."""
    biz = profile.business_income
    if not biz:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule C Expenses"

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a365d', end_color='1a365d', fill_type='solid')
    currency_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = f"Schedule C — Profit or Loss from Business"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='1a365d')

    ws.merge_cells('A2:D2')
    ws['A2'] = f"Business: {biz.business_name} | Tax Year {profile.tax_year}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)

    # Income section
    row = 4
    ws.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="Gross Receipts")
    ws.cell(row=row, column=2, value=biz.gross_receipts).number_format = currency_fmt

    # Expenses section
    row += 2
    ws.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=11)
    row += 1

    headers = ["Category", "Annual Amount", "Monthly Avg"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    exp = biz.expenses
    expense_items = [
        ("Advertising", exp.advertising),
        ("Car & Truck Expenses", exp.car_and_truck),
        ("Insurance", exp.insurance),
        ("Office Expenses", exp.office_expense),
        ("Supplies", exp.supplies),
        ("Utilities", exp.utilities),
        ("Other Expenses", exp.other),
    ]

    if biz.depreciation > 0:
        expense_items.append(("Depreciation", biz.depreciation))

    row += 1
    for cat, amt in expense_items:
        if amt > 0:
            ws.cell(row=row, column=1, value=cat).border = thin_border
            ws.cell(row=row, column=2, value=amt).number_format = currency_fmt
            ws.cell(row=row, column=2).border = thin_border
            ws.cell(row=row, column=3, value=round(amt/12, 2)).number_format = currency_fmt
            ws.cell(row=row, column=3).border = thin_border
            row += 1

    # Totals
    row += 1
    ws.cell(row=row, column=1, value="TOTAL EXPENSES").font = Font(bold=True)
    total_exp = exp.total + biz.depreciation
    ws.cell(row=row, column=2, value=total_exp).number_format = currency_fmt
    ws.cell(row=row, column=2).font = Font(bold=True)

    row += 1
    ws.cell(row=row, column=1, value="NET PROFIT").font = Font(bold=True, color='276749')
    ws.cell(row=row, column=2, value=biz.net_profit).number_format = currency_fmt
    ws.cell(row=row, column=2).font = Font(bold=True, color='276749')

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16

    wb.save(output_path)


# ===========================================================================
# Tax Document Attachments Summary
# ===========================================================================

def _generate_attachments_summary(profile, output_dir):
    """Generate a summary document listing all attachments."""
    doc = SimpleDocTemplate(
        os.path.join(output_dir, "Tax_Document_Attachments_Summary.pdf"),
        pagesize=letter,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
    )
    styles = get_styles()
    story = []

    story.append(Paragraph(
        f"Tax Document Attachments Summary — {profile.tax_year}",
        styles['DocTitle']))
    story.append(Spacer(1, 6))
    story.append(HRFlowable(width="100%", thickness=2, color=MEDIUM_BLUE))
    story.append(Spacer(1, 12))

    story.append(Paragraph(
        f"Taxpayer: <b>{profile.primary_first} {profile.primary_last}</b>"
        + (f" & <b>{profile.spouse_first} {profile.spouse_last}</b>"
           if profile.spouse_first else ""),
        styles['FieldValue']))
    story.append(Spacer(1, 12))

    # List all documents
    data = [["#", "Document Type", "Description", "Source"]]
    doc_num = 1

    for w2 in profile.w2_incomes:
        data.append([str(doc_num), "Form W-2",
                     f"Wages: {format_currency_int(w2.wages)}", w2.employer_name])
        doc_num += 1

    for ii in profile.interest_incomes:
        data.append([str(doc_num), "Form 1099-INT",
                     f"Interest: {format_currency_int(ii.amount)}", ii.payer_name])
        doc_num += 1

    for di in profile.dividend_incomes:
        data.append([str(doc_num), "Form 1099-DIV",
                     f"Dividends: {format_currency_int(di.ordinary_dividends)}",
                     di.payer_name])
        doc_num += 1

    for m in getattr(profile, 'mortgage_interests', []):
        data.append([str(doc_num), "Form 1098",
                     f"Mortgage Int: {format_currency_int(m.interest_paid)}",
                     m.lender_name])
        doc_num += 1

    for cg in getattr(profile, 'capital_gains', []):
        data.append([str(doc_num), "Form 1099-B",
                     f"Capital Gains",
                     cg.payer_name])
        doc_num += 1

    if profile.business_income:
        data.append([str(doc_num), "Bank Statement",
                     f"Business account for {profile.tax_year}",
                     profile.business_income.business_name])
        doc_num += 1
        data.append([str(doc_num), "Invoice", "Sample client invoice",
                     profile.business_income.business_name])
        doc_num += 1
        data.append([str(doc_num), "Schedule C Ledger",
                     "Annual expense breakdown",
                     profile.business_income.business_name])
        doc_num += 1

    t = Table(data, colWidths=[0.4*inch, 1.4*inch, 2.8*inch, 2.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, 0), WHITE),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    story.append(Spacer(1, 20))
    story.append(Paragraph(
        f"<b>Total documents: {doc_num - 1}</b>", styles['FieldValue']))

    doc.build(story)
